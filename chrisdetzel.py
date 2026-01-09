import re
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from datetime import datetime
import time

# try to import known scrapers
try:
    from chrisdetzel_hbm import scrape_hbm_product
except Exception:
    scrape_hbm_product = None


def setup_headless_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=options
    )
    return driver


def _choose_url_column(df):
    # Prefer column F (index 5). If not present, try to find a column containing 'url'
    if len(df.columns) >= 6:
        return df.columns[5]
    for c in df.columns:
        if 'url' in str(c).lower():
            return c
    raise ValueError('Konnte keine URL-Spalte finden (erwarte Spalte F oder eine Spalte mit "url").')


def _find_accept_column(df):
    for c in df.columns:
        low = str(c).lower()
        if 'accept' in low or 'akzept' in low:
            return c
    return None


def _domain_from_url(url):
    try:
        netloc = urlparse(str(url)).netloc.lower()
        if netloc.startswith('www.'):
            netloc = netloc[4:]
        return netloc
    except Exception:
        return ''


def _make_new_column_name(old_name, today_str):
    # If contains '_date' replace it, elif contains YYYY-MM-DD at end replace, else append
    if '_date' in old_name:
        return old_name.replace('_date', '_' + today_str)
    m = re.search(r'_(\d{4}-\d{2}-\d{2})$', old_name)
    if m:
        return re.sub(r'_(\d{4}-\d{2}-\d{2})$', '_' + today_str, old_name)
    return f"{old_name}_{today_str}"


def process_excel(path):
    # Read sheet
    sheet = 'Option 1 - nebeneinander'
        # Die erste Zeile ist leer und die zweite Zeile enthält die Header.
        # Daher header=1 verwenden und nur die ersten 6 Zeilen lesen (inkl. Header-Zeile).
        df = pd.read_excel(path, sheet_name=sheet, engine='openpyxl', header=1, nrows=6)

    url_col = _choose_url_column(df)
    accept_col = _find_accept_column(df)

    # Determine last 4 columns to copy names from
    if len(df.columns) >= 4:
        last4 = list(df.columns[-4:])
    else:
        # fallback: take all columns
        last4 = list(df.columns)

    today_str = datetime.today().strftime('%Y-%m-%d')
    new_col_names = [_make_new_column_name(str(c), today_str) for c in last4]

    # Prepare empty series for new columns
    new_cols = {name: pd.Series([pd.NA] * len(df), index=df.index, dtype=object) for name in new_col_names}

    # Domain -> function dispatch
    dispatch = {
        'hbm-machines.com': scrape_hbm_product,
    }

    for idx, row in df.iterrows():
        url = row.get(url_col)
        if not isinstance(url, str) or not url.strip():
            continue

        # Skip if acceptance column present and falsy
        if accept_col is not None:
            val = row.get(accept_col)
            if pd.isna(val) or (isinstance(val, (str,)) and val.strip().lower() in ['', 'no', 'n', 'nein', 'false']):
                # Skip this row since baselink not accepted
                continue

        domain = _domain_from_url(url)

        scraper = None
        # try exact match and suffix match
        if domain in dispatch and dispatch[domain]:
            scraper = dispatch[domain]
        else:
            for k, v in dispatch.items():
                if k and k in domain and v:
                    scraper = v
                    break

        scraped = None
        if scraper is None:
            # no scraper available for this domain -> skip
            continue

        try:
            # call scraper; it is expected to accept a url and return a dict
            scraped = scraper(url)
        except TypeError:
            # try calling with driver if scraper expects a driver
            try:
                driver = setup_headless_driver()
                scraped = scraper(url, driver)
                driver.quit()
            except Exception:
                scraped = None
        except Exception:
            scraped = None

        if not scraped:
            continue

        # Map scraped results into new columns heuristically
        # Determine a simple mapping based on keywords in last4 column names
        for old_col, new_col in zip(last4, new_col_names):
            base = str(old_col).lower()
            val = None
            if 'price' in base:
                val = scraped.get('price_raw') or scraped.get('price') or scraped.get('price_value')
            elif 'availability' in base or 'stock' in base or 'lager' in base:
                val = scraped.get('stock_normalized') or scraped.get('availability') or scraped.get('stock_raw')
            else:
                # fallback: try the key equal to the old_col (lower)
                key_guess = str(old_col)
                candidates = [k for k in scraped.keys() if k.lower() == key_guess.lower()]
                if candidates:
                    val = scraped.get(candidates[0])
                else:
                    # fallback to stringified main fields
                    val = scraped.get('price_raw') or scraped.get('stock_raw') or str(scraped)

            new_cols[new_col].iat[idx] = val

    # Append new columns to df (they will be added at the end)
    for name, series in new_cols.items():
        # if name collides, generate unique
        final_name = name
        counter = 1
        while final_name in df.columns:
            final_name = f"{name}_{counter}"
            counter += 1
        df[final_name] = series

    # Write back replacing only the sheet, preserving the rest of the workbook
    try:
        # pandas >= 1.3 supports if_sheet_exists='replace'
        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    except TypeError:
        # older pandas: load workbook, remove sheet then write
        from openpyxl import load_workbook
        wb = load_workbook(path)
        if sheet in wb.sheetnames:
            std = wb[sheet]
            wb.remove(std)
        wb.save(path)
        with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)


if __name__ == '__main__':
    EXCEL_PATH = '/Users/gabrielhipp/Library/Mobile Documents/com~apple~CloudDocs/fiverr/chrisdetzel/Preisanalyse Wettbewerber v2.1.xlsx'
    process_excel(EXCEL_PATH)
